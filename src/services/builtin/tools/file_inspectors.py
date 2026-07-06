from __future__ import annotations

import csv
import json
import mimetypes
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, List


TEXT_EXTS = {".txt", ".md", ".sql", ".py", ".js", ".ts", ".tsx", ".jsx", ".java", ".cs", ".go", ".rs", ".sh", ".bash", ".ps1", ".yaml", ".yml", ".xml", ".html", ".css", ".toml", ".ini", ".env.example"}
CODE_EXTS = {".py", ".js", ".ts", ".tsx", ".jsx", ".java", ".cs", ".go", ".rs", ".sh", ".bash", ".ps1"}


def _base(path: Path) -> Dict[str, Any]:
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return {
        "status": "success",
        "file_type": "unknown",
        "mime_type": mime,
        "size_bytes": path.stat().st_size,
        "inspection_method": "metadata",
        "summary": "Metadata only.",
        "facts": {},
        "sample": {},
        "warnings": [],
        "truncated": False,
    }


def _infer_simple_type(values: List[str]) -> str:
    v = [x for x in values if x not in (None, "")]
    if not v:
        return "empty"
    if all(re.fullmatch(r"-?\d+", x or "") for x in v):
        return "int"
    if all(re.fullmatch(r"-?\d+(\.\d+)?", x or "") for x in v):
        return "float"
    return "string"


def inspect_text_or_code(path: Path, *, max_chars: int = 4000) -> Dict[str, Any]:
    out = _base(path)
    ext = path.suffix.lower()
    raw = path.read_text(encoding="utf-8", errors="replace")
    truncated = len(raw) > max_chars
    preview = raw[:max_chars]
    lines = raw.splitlines()

    file_type = "code" if ext in CODE_EXTS else "text"
    out.update({
        "file_type": file_type,
        "inspection_method": "text_preview",
        "summary": f"{file_type} file inspected.",
        "facts": {
            "extension": ext,
            "line_count": len(lines),
            "char_count": len(raw),
            "language": ext.lstrip("."),
        },
        "sample": {
            "preview": preview,
            "first_lines": lines[:20],
        },
        "truncated": truncated,
    })
    if truncated:
        out["warnings"].append("Preview truncated for safety.")

    if file_type == "code":
        imports = re.findall(r"^(?:from\s+([\w\.]+)\s+import|import\s+([\w\.]+))", raw, flags=re.MULTILINE)
        flat_imports = [a or b for a, b in imports][:30]
        out["facts"]["imports"] = flat_imports
    if ext == ".sql":
        tables = re.findall(r"\b(?:from|join|into|update)\s+([a-zA-Z0-9_\.]+)", raw, flags=re.IGNORECASE)
        out["facts"]["table_references"] = sorted(set(tables))[:50]
    return out


def inspect_json(path: Path, *, max_depth: int = 3, max_chars: int = 4000) -> Dict[str, Any]:
    out = _base(path)
    raw = path.read_text(encoding="utf-8", errors="replace")
    obj = json.loads(raw)

    def shape(x: Any, depth: int) -> Any:
        if depth <= 0:
            return type(x).__name__
        if isinstance(x, dict):
            return {k: shape(v, depth - 1) for k, v in list(x.items())[:20]}
        if isinstance(x, list):
            if not x:
                return []
            return [shape(x[0], depth - 1)]
        return type(x).__name__

    root_type = "dict" if isinstance(obj, dict) else "list" if isinstance(obj, list) else type(obj).__name__
    top_keys = list(obj.keys())[:100] if isinstance(obj, dict) else []
    array_lengths = {k: len(v) for k, v in obj.items() if isinstance(v, list)} if isinstance(obj, dict) else {"root": len(obj)} if isinstance(obj, list) else {}

    out.update({
        "file_type": "json",
        "inspection_method": "json_parse",
        "summary": "JSON inspected.",
        "facts": {
            "root_type": root_type,
            "top_level_keys": top_keys,
            "array_lengths": array_lengths,
        },
        "sample": {
            "structure": shape(obj, max_depth),
            "preview": raw[:max_chars],
        },
        "truncated": len(raw) > max_chars,
    })
    if out["truncated"]:
        out["warnings"].append("JSON preview truncated for safety.")
    return out


def inspect_jsonl(path: Path, *, max_rows: int = 50) -> Dict[str, Any]:
    out = _base(path)
    valid = 0
    invalid = 0
    sampled = 0
    keyset = set()
    sample_records: List[Any] = []
    with path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            if sampled >= max_rows:
                break
            sampled += 1
            s = line.strip()
            if not s:
                continue
            try:
                rec = json.loads(s)
                valid += 1
                if isinstance(rec, dict):
                    keyset.update(list(rec.keys())[:50])
                if len(sample_records) < 5:
                    sample_records.append(rec)
            except Exception:
                invalid += 1

    out.update({
        "file_type": "jsonl",
        "inspection_method": "jsonl_sample",
        "summary": "JSONL sampled.",
        "facts": {
            "sampled_line_count": sampled,
            "valid_sample_count": valid,
            "invalid_sample_count": invalid,
            "sample_keys": sorted(keyset)[:100],
        },
        "sample": {"records": sample_records},
        "truncated": True,
    })
    out["warnings"].append("JSONL analysis based on sampled lines.")
    return out


def inspect_notebook(path: Path, *, max_cells: int = 20) -> Dict[str, Any]:
    out = inspect_json(path, max_depth=2)
    obj = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    cells = obj.get("cells") if isinstance(obj, dict) else []
    if not isinstance(cells, list):
        cells = []
    code_cells = [c for c in cells if isinstance(c, dict) and c.get("cell_type") == "code"]
    md_cells = [c for c in cells if isinstance(c, dict) and c.get("cell_type") == "markdown"]
    operations = ["read", "join", "filter", "select", "withColumn", "write", "saveAsTable", "merge", "spark.sql"]
    found_ops = set()
    imports = set()
    tables = set()
    writes = set()
    previews = []
    for c in cells[:max_cells]:
        src = "".join(c.get("source") or []) if isinstance(c, dict) else ""
        for op in operations:
            if op.lower() in src.lower():
                found_ops.add(op)
        for a, b in re.findall(r"^(?:from\s+([\w\.]+)\s+import|import\s+([\w\.]+))", src, re.MULTILINE):
            imports.add(a or b)
        for t in re.findall(r"\b(?:from|join|table|saveAsTable|insert into)\s+([a-zA-Z0-9_\.]+)", src, re.IGNORECASE):
            tables.add(t)
        for w in re.findall(r"\b(?:saveAsTable|insert into|overwrite|write\.[a-z]+\(['\"]([^'\"]+))", src, re.IGNORECASE):
            writes.add(w)
        previews.append({"cell_type": c.get("cell_type"), "preview": src[:300]})
    out.update({
        "file_type": "notebook",
        "inspection_method": "notebook_summary",
        "summary": "Notebook inspected.",
        "facts": {
            "notebook_type": obj.get("nbformat"),
            "cell_count": len(cells),
            "code_cell_count": len(code_cells),
            "markdown_cell_count": len(md_cells),
            "imports": sorted(imports)[:50],
            "referenced_tables": sorted(tables)[:50],
            "write_targets": sorted(writes)[:50],
            "detected_operations": sorted(found_ops),
        },
        "sample": {"cell_previews": previews[:max_cells]},
        "truncated": len(cells) > max_cells,
    })
    return out


def inspect_csv(path: Path, *, max_rows: int = 20) -> Dict[str, Any]:
    out = _base(path)
    ext = path.suffix.lower()
    delim = "\t" if ext == ".tsv" else ","
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        sample_text = f.read(4096)
    try:
        sniffed = csv.Sniffer().sniff(sample_text, delimiters=",\t;|")
        delim = sniffed.delimiter
    except Exception:
        pass

    columns: List[str] = []
    sample_rows: List[Dict[str, Any]] = []
    col_values: Dict[str, List[str]] = {}
    null_counts: Dict[str, int] = {}
    row_count = 0
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(f, delimiter=delim)
        columns = reader.fieldnames or []
        for col in columns:
            col_values[col] = []
            null_counts[col] = 0
        for row in reader:
            row_count += 1
            if len(sample_rows) < max_rows:
                sample_rows.append(row)
            for c in columns:
                val = row.get(c)
                if val in (None, ""):
                    null_counts[c] += 1
                if len(col_values[c]) < max_rows and val is not None:
                    col_values[c].append(val)

    types = {c: _infer_simple_type(col_values.get(c, [])) for c in columns}
    out.update({
        "file_type": "csv",
        "inspection_method": "csv_sample",
        "summary": "Delimited file profiled.",
        "facts": {
            "delimiter": delim,
            "columns": columns,
            "row_count": row_count,
            "inferred_types": types,
            "null_or_empty_counts": null_counts,
        },
        "sample": {"rows": sample_rows[:max_rows]},
        "truncated": row_count > max_rows,
    })
    if row_count > max_rows:
        out["warnings"].append("Only sampled rows are included.")
    return out


def inspect_archive(path: Path, *, max_entries: int = 50) -> Dict[str, Any]:
    out = _base(path)
    entries = []
    with zipfile.ZipFile(path, "r") as z:
        infos = z.infolist()
        total_uncompressed = sum(i.file_size for i in infos)
        for i in infos[:max_entries]:
            entries.append({"path": i.filename, "size": i.file_size, "compressed_size": i.compress_size})
    out.update({
        "file_type": "archive",
        "inspection_method": "zip_metadata",
        "summary": "Archive entries listed without extraction.",
        "facts": {"file_count": len(infos), "total_uncompressed_size": total_uncompressed},
        "sample": {"entries": entries},
        "truncated": len(infos) > max_entries,
    })
    return out


def inspect_parquet(path: Path, *, max_rows: int = 10) -> Dict[str, Any]:
    out = _base(path)
    try:
        import pyarrow.parquet as pq
    except Exception:
        out.update({"status": "partial", "file_type": "parquet", "inspection_method": "metadata_only", "summary": "pyarrow not installed."})
        out["warnings"].append("pyarrow not installed; parquet inspection limited.")
        return out
    pf = pq.ParquetFile(path)
    schema = [str(x) for x in pf.schema]
    cols = pf.schema.names
    rows = pf.metadata.num_rows if pf.metadata else None
    sample = []
    try:
        table = pq.read_table(path, columns=cols[: min(10, len(cols))])
        sample = table.slice(0, max_rows).to_pylist()
    except Exception:
        out["warnings"].append("Could not load parquet sample rows.")
    out.update({"file_type": "parquet", "inspection_method": "parquet_metadata", "summary": "Parquet metadata inspected.", "facts": {"schema": schema[:100], "columns": cols, "row_count": rows}, "sample": {"rows": sample}})
    return out


def inspect_excel(path: Path, *, max_rows: int = 10) -> Dict[str, Any]:
    out = _base(path)
    ext = path.suffix.lower()
    if ext != ".xlsx":
        out.update({"status": "partial", "file_type": "excel", "summary": "Only .xlsx supported by openpyxl path."})
        out["warnings"].append(".xls parsing not supported without extra dependencies.")
        return out
    try:
        from openpyxl import load_workbook
    except Exception:
        out.update({"status": "partial", "file_type": "excel", "summary": "openpyxl not installed."})
        out["warnings"].append("openpyxl not installed; excel inspection limited.")
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    sample_sheets = []
    for name in sheet_names[:5]:
        ws = wb[name]
        rows = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
            rows.append(list(row))
            if idx >= max_rows:
                break
        sample_sheets.append({"name": name, "dimensions": ws.calculate_dimension(), "sample_rows": rows, "header": rows[0] if rows else []})
    out.update({"file_type": "excel", "inspection_method": "openpyxl_read_only", "summary": "Excel workbook inspected.", "facts": {"sheet_count": len(sheet_names), "sheet_names": sheet_names}, "sample": {"sheets": sample_sheets}})
    return out


def inspect_pdf(path: Path, *, max_chars: int = 2000) -> Dict[str, Any]:
    out = _base(path)
    try:
        from pypdf import PdfReader
    except Exception:
        out.update({"status": "partial", "file_type": "pdf", "summary": "PDF parser dependency missing."})
        out["warnings"].append("pypdf not installed; pdf inspection limited.")
        return out
    reader = PdfReader(str(path))
    meta = dict(reader.metadata or {})
    text_preview = ""
    for p in reader.pages[:3]:
        try:
            text_preview += (p.extract_text() or "") + "\n"
        except Exception:
            pass
    text_preview = text_preview[:max_chars]
    out.update({"file_type": "pdf", "inspection_method": "pypdf_preview", "summary": "PDF metadata and preview extracted.", "facts": {"page_count": len(reader.pages), "metadata": {str(k): str(v)[:200] for k, v in meta.items()}}, "sample": {"text_preview": text_preview}, "truncated": True})
    return out


def inspect_binary_or_unknown(path: Path) -> Dict[str, Any]:
    out = _base(path)
    ext = path.suffix.lower()
    out.update({
        "status": "partial",
        "file_type": "binary" if out["mime_type"].startswith("application/") else "unknown",
        "inspection_method": "metadata_only",
        "summary": "Binary/unknown file. Metadata only.",
        "facts": {"extension": ext, "inspection_recommendation": "Use specialized parser tool or convert to text format."},
    })
    return out


def dispatch_inspect(path: Path, *, inspection_goal: str = "", max_chars: int = 4000, max_rows: int = 20, max_cells: int = 20) -> Dict[str, Any]:
    ext = path.name.lower()
    suffix = path.suffix.lower()
    if ext.endswith(".env.example"):
        return inspect_text_or_code(path, max_chars=max_chars)
    if suffix in TEXT_EXTS:
        return inspect_text_or_code(path, max_chars=max_chars)
    if suffix == ".json":
        return inspect_json(path, max_chars=max_chars)
    if suffix in {".jsonl", ".ndjson"}:
        return inspect_jsonl(path, max_rows=max_rows)
    if suffix == ".ipynb":
        return inspect_notebook(path, max_cells=max_cells)
    if suffix in {".csv", ".tsv"}:
        return inspect_csv(path, max_rows=max_rows)
    if suffix == ".zip":
        return inspect_archive(path, max_entries=max_rows)
    if suffix == ".parquet":
        return inspect_parquet(path, max_rows=max_rows)
    if suffix in {".xlsx", ".xls"}:
        return inspect_excel(path, max_rows=max_rows)
    if suffix == ".pdf":
        return inspect_pdf(path, max_chars=max_chars)
    return inspect_binary_or_unknown(path)
